import os
import json
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import shutil

# Rutas de configuración
PEDIDOS_DIR = r'C:\Users\Sebastian Escobar P\Documents\Sebastian\Oficina\Antigravity\pedidos_recibidos'
EXCEL_PRECIOS = r'c:\Users\Sebastian Escobar P\Documents\Sebastian\Oficina\Antigravity\inputs\lista_precios.xlsx'
SALIDA_DIR = r'C:\Users\Sebastian Escobar P\Documents\Sebastian\Oficina\Antigravity\compras_proveedores'
PLANTILLA_PO = r'C:\Users\Sebastian Escobar P\Documents\Sebastian\Oficina\Sescop LLC\Ordenes de compra\Plantilla orden de compra.xlsx'

def consolidar_pedidos():
    # 1. Asegurar que las carpetas existan
    if not os.path.exists(PEDIDOS_DIR):
        print(f"Creando carpeta de pedidos en: {PEDIDOS_DIR}")
        os.makedirs(PEDIDOS_DIR)
        print("Pon los archivos .json que te envíe la vendedora en esa carpeta y vuelve a correr este programa.")
        return

    os.makedirs(SALIDA_DIR, exist_ok=True)

    if not os.path.exists(PLANTILLA_PO):
        print(f"ERROR: No se encontró la plantilla en: {PLANTILLA_PO}")
        return

    # 2. Leer todos los archivos JSON en la carpeta de pedidos
    archivos = [f for f in os.listdir(PEDIDOS_DIR) if f.endswith('.json')]
    if not archivos:
        print(f"No se encontraron archivos .json en {PEDIDOS_DIR}")
        return

    print(f"Procesando {len(archivos)} pedidos...")
    
    productos_totales = {} # {sku: cantidad}
    
    for archivo in archivos:
        path = os.path.join(PEDIDOS_DIR, archivo)
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for item in data.get('items', []):
                    sku = item.get('sku')
                    qty = item.get('quantity', 0)
                    productos_totales[sku] = productos_totales.get(sku, 0) + qty
        except Exception as e:
            print(f"Error leyendo {archivo}: {e}")

    # 3. Cargar el Excel maestro para obtener Proveedor y Costo
    try:
        master_df = pd.read_excel(EXCEL_PRECIOS)
        # Limpiar SKU y columnas
        master_df.columns = [str(c).strip() for c in master_df.columns]
        master_df['SKU'] = master_df['SKU'].astype(str).str.strip()
    except Exception as e:
        print(f"Error cargando el Excel maestro: {e}")
        return

    # 4. Cruzar datos y organizar por Proveedor
    proveedores_data = {} # {proveedor: [items]}

    for sku, qty in productos_totales.items():
        row = master_df[master_df['SKU'] == sku]
        if row.empty:
            print(f"Advertencia: SKU {sku} no encontrado en el Excel maestro.")
            continue
        
        def clean_price(val):
            try:
                if pd.isna(val): return 0.0
                if isinstance(val, (int, float)): return float(val)
                s = str(val).replace('$', '').replace(',', '').strip()
                if not s or any(x in s for x in ['+', '%', 'PRECIO']): return 0.0
                return float(s)
            except:
                return 0.0

        row_data = row.iloc[0]
        proveedor = str(row_data.get('Proveedor', 'SIN PROVEEDOR')).strip()
        nombre = str(row_data.get('PRODUCTO', 'N/A')).strip()
        costo = clean_price(row_data.get('Costo', 0))
        
        item_final = {
            'SKU': sku,
            'Producto': nombre,
            'Cantidad': qty,
            'Costo': costo
        }
        
        if proveedor not in proveedores_data:
            proveedores_data[proveedor] = []
        proveedores_data[proveedor].append(item_final)

    # 5. Generar un Excel usando la plantilla por cada proveedor
    fecha_hoy = datetime.now().strftime("%Y-%m-%d_%H-%M")
    
    for proveedor, items in proveedores_data.items():
        # Crear nombre de archivo de salida
        prov_file = f"COMPRA_{proveedor.replace(' ', '_')}_{fecha_hoy}.xlsx"
        path_salida = os.path.join(SALIDA_DIR, prov_file)
        
        # Copiar plantilla
        shutil.copy(PLANTILLA_PO, path_salida)
        
        # Cargar el libro copiado
        wb = load_workbook(path_salida)
        ws = wb.active # Asumimos que la plantilla está en la primera hoja activa
        
        # Empezar a escribir desde la fila 6
        fila_inicio = 6
        gran_total = 0
        ultima_fila = fila_inicio
        
        for i, item in enumerate(items):
            current_row = fila_inicio + i
            ultima_fila = current_row
            
            qty = item['Cantidad']
            costo = item['Costo']
            subtotal = qty * costo
            gran_total += subtotal
            
            ws[f'A{current_row}'] = qty
            ws[f'B{current_row}'] = item['SKU']
            ws[f'C{current_row}'] = item['Producto']
            ws[f'D{current_row}'] = costo
            ws[f'E{current_row}'] = subtotal
            
            # Aplicar formato de moneda
            ws[f'D{current_row}'].number_format = '$#,##0.00'
            ws[f'E{current_row}'].number_format = '$#,##0.00'

        # Escribir el GRAN TOTAL al final
        fila_total = ultima_fila + 2
        ws[f'D{fila_total}'] = "TOTAL ORDEN:"
        ws[f'E{fila_total}'] = gran_total
        
        # Estilo para el total
        from openpyxl.styles import Font, Alignment
        ws[f'D{fila_total}'].font = Font(bold=True)
        ws[f'E{fila_total}'].font = Font(bold=True)
        ws[f'E{fila_total}'].number_format = '$#,##0.00'
        ws[f'D{fila_total}'].alignment = Alignment(horizontal='right')

        wb.save(path_salida)
        print(f"Generada orden para {proveedor}: {prov_file} | Total: ${gran_total:,.2f}")

    print("\n¡Proceso terminado con éxito!")
    print(f"Revisa las órdenes en: {SALIDA_DIR}")

if __name__ == "__main__":
    consolidar_pedidos()
